import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# Configuration
INPUT_FILE = r"WorkPlan/ACMS-HIV CHASAC WorkPlan-FY26-COP25 Updated 18.11.25.xlsx"
OUTPUT_FILE = "Full_Workplan_Tracker.xlsx"
SHEET_NAME = '4. ACMS WorkPlan detail v1'

def extract_all_tasks():
    if not os.path.exists(INPUT_FILE):
        print(f"Error: Input file not found at {INPUT_FILE}")
        return

    print(f"Reading {INPUT_FILE}...")
    print(f"Reading {INPUT_FILE}...")
    try:
        from tqdm import tqdm
        
        with tqdm(total=8, desc="Overall Progress", unit="step") as pbar:
            # Read with header=1
            df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME, header=1)
            pbar.update(1)
            pbar.set_description("Forward Filling")
            
            # 1. Forward Fill Parent Columns
            # In Excel workplans, often the "Activities" or "Program Area" is only listed on the first row of a group.
            # We need to propagate these down to the sub-activities.
            cols_to_ffill = ['Activities', 'Code Sub -activities', 'Program Area', 'Sub-Activity Category']
            
            # Only ffill if the column exists
            for col in cols_to_ffill:
                if col in df.columns:
                    df[col] = df[col].ffill()
            pbar.update(1)
            pbar.set_description("Filtering Rows")
            
            # 2. Filter out completely empty rows (if any)
            # Check if 'ACMS Sub-Activities' is not null, as that seems to be the main task line
            # Or just drop rows where all columns are NaN
            df = df.dropna(how='all')
            pbar.update(1)
            pbar.set_description("Selecting Columns")
            
            # 3. Select relevant columns
            cols_to_keep = [
                'Activities', 
                'Code Sub -activities',
                'ACMS Sub-Activities', 
                'Program Area', 
                'Level of Activity Implementation (Above-Site, Site-Level)', 
                'Sub-Activity Category', 
                'Oct -Dec 2025', 
                'Jan - Mar 2026', 
                'Outputs', 
                'Output Indicators'
            ]
            
            existing_cols = [c for c in cols_to_keep if c in df.columns]
            tasks = df[existing_cols].copy()
            pbar.update(1)
            pbar.set_description("Cleaning Data")
            
            # 4. Clean remaining NaNs
            # Replace NaN with empty string for cleaner display
            tasks = tasks.fillna('')
            
            # 5. Add Tracking Columns
            tasks.insert(0, 'ID', range(1, len(tasks) + 1))
            tasks['Status'] = 'Pending'
            tasks['Progress (%)'] = 0
            tasks['Comments'] = ''
            tasks['Assigned To'] = '' # User can fill this in
            pbar.update(1)
            
            print(f"\nExtracted {len(tasks)} tasks.")
            pbar.set_description("Saving Excel")
            
            # Save to new Excel
            tasks.to_excel(OUTPUT_FILE, index=False, sheet_name='All_Tasks')
            pbar.update(1)
            pbar.set_description("Formatting Excel")
            
            # Formatting
            wb = load_workbook(OUTPUT_FILE)
            ws = wb['All_Tasks']
            
            # Header formatting
            header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid") # Different color for full tracker
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                
            # Auto-adjust column widths
            # Use a nested progress bar for columns as this can be slow
            for col in tqdm(list(ws.columns), desc="Adjusting Widths", leave=False):
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        val_len = len(str(cell.value))
                        if val_len > max_length:
                            max_length = val_len
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60) 
                ws.column_dimensions[column].width = adjusted_width
            
            pbar.update(1)
            pbar.set_description("Finalizing")
                
            wb.save(OUTPUT_FILE)
            pbar.update(1)
            pbar.set_description("Complete")
            
        print("Extraction complete!")
        
    except Exception as e:
        print(f"An error occurred: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    extract_all_tasks()
