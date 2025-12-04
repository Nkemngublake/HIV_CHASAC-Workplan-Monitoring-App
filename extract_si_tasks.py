import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# Configuration
INPUT_FILE = r"WorkPlan/ACMS-HIV CHASAC WorkPlan-FY26-COP25 Updated 18.11.25.xlsx"
OUTPUT_FILE = "SI_Manager_Tracker.xlsx"
SHEET_NAME = '4. ACMS WorkPlan detail v1'

def extract_tasks():
    if not os.path.exists(INPUT_FILE):
        print(f"Error: Input file not found at {INPUT_FILE}")
        return

    print(f"Reading {INPUT_FILE}...")
    try:
        from tqdm import tqdm
        
        with tqdm(total=6, desc="SI Extraction Progress", unit="step") as pbar:
            # Read with header=1 as determined in analysis
            df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME, header=1)
            pbar.update(1)
            pbar.set_description("Filtering Data")
            
            # Filter criteria
            # 1. Program Area contains "SI", "Strategic Information", "M&E", "Data"
            # 2. Activities contains "Objective 7"
            # 3. Sub-Activities contains "Data", "SI ", "M&E", "DQA"
            
            # Create masks
            mask_program_area = df['Program Area'].astype(str).str.contains('SI|Strategic Information|M&E|Data|Cross-Cutting\(SI/H\)', case=False, na=False)
            mask_objective_7 = df['Activities'].astype(str).str.contains('Objective 7', case=False, na=False)
            mask_sub_activities = df['ACMS Sub-Activities'].astype(str).str.contains('Data|SI |Strategic Information|M&E|DQA', case=False, na=False)
            
            # Combine masks (OR logic)
            final_mask = mask_program_area | mask_objective_7 | mask_sub_activities
            
            si_tasks = df[final_mask].copy()
            
            print(f"\nFound {len(si_tasks)} SI-related tasks.")
            pbar.update(1)
            pbar.set_description("Selecting Columns")
            
            # Select relevant columns
            # We want to keep context but simplify if possible. 
            # Let's keep: Activities, ACMS Sub-Activities, Program Area, Level, Category, Timeline (Oct-Dec, Jan-Mar), Outputs
            
            cols_to_keep = [
                'Activities', 
                'ACMS Sub-Activities', 
                'Program Area', 
                'Level of Activity Implementation (Above-Site, Site-Level)', 
                'Sub-Activity Category', 
                'Oct -Dec 2025', 
                'Jan - Mar 2026', 
                'Outputs', 
                'Output Indicators'
            ]
            
            # Filter columns that actually exist
            existing_cols = [c for c in cols_to_keep if c in si_tasks.columns]
            si_tasks = si_tasks[existing_cols]
            pbar.update(1)
            pbar.set_description("Adding Tracking Info")
            
            # Add Tracking Columns
            si_tasks.insert(0, 'ID', range(1, len(si_tasks) + 1))
            si_tasks['Status'] = 'Pending' # Pending, In Progress, Completed, Delayed
            si_tasks['Progress (%)'] = 0
            si_tasks['Comments'] = ''
            si_tasks['Assigned To'] = 'SI Manager'
            pbar.update(1)
            pbar.set_description("Saving Excel")
            
            # Save to new Excel
            si_tasks.to_excel(OUTPUT_FILE, index=False, sheet_name='SI_Tasks')
            pbar.update(1)
            pbar.set_description("Formatting Excel")
            
            # Formatting
            wb = load_workbook(OUTPUT_FILE)
            ws = wb['SI_Tasks']
            
            # Header formatting
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                
            # Auto-adjust column widths (approximate)
            for col in tqdm(list(ws.columns), desc="Adjusting Widths", leave=False):
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50) # Cap width at 50
                ws.column_dimensions[column].width = adjusted_width
                
            wb.save(OUTPUT_FILE)
            pbar.update(1)
            pbar.set_description("Complete")
            
        print("Extraction complete!")
        
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    extract_tasks()
